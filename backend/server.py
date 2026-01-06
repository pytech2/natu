from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, Form, Query, Header, Request
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import bcrypt
import jwt
from bson import ObjectId
import aiofiles
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import json
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.pdfgen import canvas
from PIL import Image as PILImage, ImageDraw, ImageFont
import tempfile
import fitz  # PyMuPDF for PDF processing
import re
import math

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Configuration
JWT_SECRET = os.environ.get('JWT_SECRET', 'nstu-property-tax-secret-key-2025')
JWT_ALGORITHM = "HS256"

# Create uploads directory
UPLOAD_DIR = ROOT_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

# Create the main app
app = FastAPI(title="NSTU Property Tax Manager")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# ============== MODELS ==============

# Role options: ADMIN, SURVEYOR, SUPERVISOR, MC_OFFICER
class UserCreate(BaseModel):
    username: str
    password: str
    name: str
    role: str = "SURVEYOR"  # ADMIN, SURVEYOR, SUPERVISOR, MC_OFFICER
    assigned_area: Optional[str] = None

class UserLogin(BaseModel):
    username: str
    password: str

class UserResponse(BaseModel):
    id: str
    username: str
    name: str
    role: str
    assigned_area: Optional[str] = None
    created_at: str

class TokenResponse(BaseModel):
    token: str
    user: UserResponse

class DatasetBatchCreate(BaseModel):
    name: str

class DatasetBatchResponse(BaseModel):
    id: str
    name: str
    uploaded_by: str
    uploaded_at: str
    status: str
    total_records: int

class PropertyResponse(BaseModel):
    id: str
    batch_id: str
    property_id: str
    owner_name: str
    mobile: str
    address: str
    total_area: Optional[str] = None
    amount: Optional[str] = None
    ward: Optional[str] = None
    assigned_employee_id: Optional[str] = None
    assigned_employee_name: Optional[str] = None
    status: str
    created_at: str

class AssignmentRequest(BaseModel):
    property_ids: List[str]
    employee_id: str

class BulkAssignmentRequest(BaseModel):
    area: str
    employee_id: str

class SubmissionApproval(BaseModel):
    submission_id: str
    action: str  # APPROVE or REJECT
    remarks: Optional[str] = None

class DashboardStats(BaseModel):
    total_properties: int
    completed: int
    pending: int
    in_progress: int
    rejected: int
    employees: int
    batches: int
    today_completed: int
    today_wards: int

class EmployeeProgress(BaseModel):
    employee_id: str
    employee_name: str
    role: str
    total_assigned: int
    completed: int
    pending: int
    today_completed: int
    overall_completed: int

# ============== HELPER FUNCTIONS ==============

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode(), hashed.encode())

def create_token(user_id: str, role: str) -> str:
    payload = {
        "user_id": user_id,
        "role": role,
        "exp": datetime.now(timezone.utc).timestamp() + 86400 * 7  # 7 days
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(authorization: str = Header(None)):
    if not authorization:
        raise HTTPException(status_code=401, detail="Token required")
    try:
        token = authorization
        if token.startswith("Bearer "):
            token = token[7:]
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user = await db.users.find_one({"id": payload["user_id"]}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

def get_today_start():
    """Get the start of today in UTC"""
    now = datetime.now(timezone.utc)
    return now.replace(hour=0, minute=0, second=0, microsecond=0)

# ============== AUTH ROUTES ==============

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(data: UserLogin):
    user = await db.users.find_one({"username": data.username}, {"_id": 0})
    if not user or not verify_password(data.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    token = create_token(user["id"], user["role"])
    return {
        "token": token,
        "user": {
            "id": user["id"],
            "username": user["username"],
            "name": user["name"],
            "role": user["role"],
            "assigned_area": user.get("assigned_area"),
            "created_at": user["created_at"]
        }
    }

@api_router.get("/auth/me")
async def get_me(current_user: dict = Depends(get_current_user)):
    return {
        "id": current_user["id"],
        "username": current_user["username"],
        "name": current_user["name"],
        "role": current_user["role"],
        "assigned_area": current_user.get("assigned_area"),
        "created_at": current_user["created_at"]
    }

# ============== ADMIN USER ROUTES ==============

@api_router.post("/admin/users", response_model=UserResponse)
async def create_user(data: UserCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "ADMIN":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    existing = await db.users.find_one({"username": data.username})
    if existing:
        raise HTTPException(status_code=400, detail="Username already exists")
    
    user_doc = {
        "id": str(uuid.uuid4()),
        "username": data.username,
        "password_hash": hash_password(data.password),
        "name": data.name,
        "role": data.role,
        "assigned_area": data.assigned_area,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user_doc)
    
    return {
        "id": user_doc["id"],
        "username": user_doc["username"],
        "name": user_doc["name"],
        "role": user_doc["role"],
        "assigned_area": user_doc["assigned_area"],
        "created_at": user_doc["created_at"]
    }

@api_router.get("/admin/users", response_model=List[UserResponse])
async def list_users(current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "ADMIN":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(1000)
    return users

@api_router.delete("/admin/users/{user_id}")
async def delete_user(user_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "ADMIN":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"message": "User deleted"}

# ============== BATCH UPLOAD ROUTES ==============

@api_router.post("/admin/batch/upload")
async def upload_batch(
    file: UploadFile = File(...),
    batch_name: str = Form(...),
    authorization: str = Form(...)
):
    current_user = await get_current_user(authorization)
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Read file content
    content = await file.read()
    filename = file.filename.lower()
    
    properties = []
    
    # Check file type and parse accordingly
    if filename.endswith('.xlsx') or filename.endswith('.xls'):
        # Parse Excel file using openpyxl
        import openpyxl
        from io import BytesIO
        
        workbook = openpyxl.load_workbook(BytesIO(content), data_only=True)
        sheet = workbook.active
        
        # Get headers from first row
        headers = []
        for cell in sheet[1]:
            headers.append(str(cell.value).strip() if cell.value else "")
        
        # Create header mapping (case-insensitive)
        header_map = {h.lower(): i for i, h in enumerate(headers)}
        
        # Parse data rows starting from row 2
        serial_num = 1
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):  # Skip empty rows
                continue
            
            # Get values using header mapping
            def get_val(keys):
                for k in keys:
                    idx = header_map.get(k.lower())
                    if idx is not None and idx < len(row) and row[idx]:
                        return str(row[idx]).strip()
                return ""
            
            prop = {
                "id": str(uuid.uuid4()),
                "serial_number": serial_num,
                "property_id": get_val(["Property Id", "property_id", "PropertyID"]) or str(uuid.uuid4())[:8].upper(),
                "old_property_id": get_val(["Old Property Id", "old_property_id", "OldPropertyId"]),
                "owner_name": get_val(["Owner Name", "owner_name", "OwnerName"]) or "Unknown",
                "mobile": get_val(["Mobile", "mobile", "Mobile No", "Phone"]),
                "address": get_val(["Plot Address", "Address", "address", "plot_address"]),
                "colony": get_val(["Colony", "colony", "Area", "area"]),
                "ward": get_val(["Colony", "Ward", "ward", "area"]),
                "latitude": None,
                "longitude": None,
                "total_area": get_val(["Total Area (SqYard)", "Total Area", "total_area", "Area"]),
                "category": get_val(["Category", "category"]),
                "amount": get_val(["Outstanding", "Total Outstanding", "Amount", "amount"]) or "0",
                "financial_year": get_val(["Financial Year", "financial_year"]) or "2025-2026",
                "assigned_employee_id": None,
                "assigned_employee_name": None,
                "status": "Pending",
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            
            # Parse latitude/longitude if available
            lat_str = get_val(["Latitude", "latitude", "Lat"])
            lng_str = get_val(["Longitude", "longitude", "Long", "Lng"])
            if lat_str:
                try:
                    prop["latitude"] = float(lat_str)
                except:
                    pass
            if lng_str:
                try:
                    prop["longitude"] = float(lng_str)
                except:
                    pass
            
            properties.append(prop)
            serial_num += 1
    else:
        # Parse CSV file
        content_str = content.decode('utf-8')
        reader = csv.DictReader(io.StringIO(content_str))
        
        serial_num = 1
        for row in reader:
            prop = {
                "id": str(uuid.uuid4()),
                "serial_number": serial_num,
                "property_id": row.get("property_id") or row.get("Property Id") or row.get("PropertyID") or str(uuid.uuid4())[:8].upper(),
                "old_property_id": row.get("old_property_id") or row.get("Old Property Id") or "",
                "owner_name": row.get("owner_name") or row.get("Owner Name") or row.get("OwnerName") or "Unknown",
                "mobile": row.get("mobile") or row.get("Mobile") or row.get("Mobile No") or "",
                "address": row.get("address") or row.get("Address") or row.get("Plot Address") or row.get("plot_address") or "",
                "colony": row.get("Colony") or row.get("colony") or row.get("Area") or "",
                "ward": row.get("ward") or row.get("Ward") or row.get("Colony") or row.get("area") or "",
                "latitude": None,
                "longitude": None,
                "total_area": row.get("total_area") or row.get("Total Area") or row.get("Total Area (SqYard)") or "",
                "category": row.get("Category") or row.get("category") or "",
                "amount": row.get("amount") or row.get("Amount") or row.get("Outstanding") or "0",
                "financial_year": row.get("Financial Year") or row.get("financial_year") or "2025-2026",
                "assigned_employee_id": None,
                "assigned_employee_name": None,
                "status": "Pending",
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            
            # Parse latitude/longitude
            lat_str = row.get("Latitude") or row.get("latitude")
            lng_str = row.get("Longitude") or row.get("longitude")
            if lat_str:
                try:
                    prop["latitude"] = float(lat_str)
                except:
                    pass
            if lng_str:
                try:
                    prop["longitude"] = float(lng_str)
                except:
                    pass
            
            properties.append(prop)
            serial_num += 1
    
    if not properties:
        raise HTTPException(status_code=400, detail="No valid properties found in file")
    
    # Create batch
    batch_doc = {
        "id": str(uuid.uuid4()),
        "name": batch_name,
        "uploaded_by": current_user["id"],
        "uploaded_at": datetime.now(timezone.utc).isoformat(),
        "status": "ACTIVE",
        "total_records": len(properties)
    }
    await db.batches.insert_one(batch_doc)
    
    # Add batch_id to properties and insert
    for prop in properties:
        prop["batch_id"] = batch_doc["id"]
    
    if properties:
        await db.properties.insert_many(properties)
    
    return {
        "batch_id": batch_doc["id"],
        "name": batch_doc["name"],
        "total_records": len(properties),
        "message": f"Successfully uploaded {len(properties)} properties"
    }

@api_router.get("/admin/batches", response_model=List[DatasetBatchResponse])
async def list_batches(current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "ADMIN":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    batches = await db.batches.find({"status": {"$ne": "DELETED"}}, {"_id": 0}).to_list(100)
    return batches

@api_router.post("/admin/batch/{batch_id}/archive")
async def archive_batch(batch_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "ADMIN":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    result = await db.batches.update_one(
        {"id": batch_id},
        {"$set": {"status": "ARCHIVED"}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Batch not found")
    return {"message": "Batch archived"}

@api_router.delete("/admin/batch/{batch_id}")
async def delete_batch(batch_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "ADMIN":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    await db.properties.delete_many({"batch_id": batch_id})
    await db.submissions.delete_many({"batch_id": batch_id})
    await db.batches.delete_one({"id": batch_id})
    
    return {"message": "Batch and all related data deleted"}

# ============== ROLE DEFINITIONS ==============
# Roles with admin-level access (can modify data)
ADMIN_ROLES = ["ADMIN", "SUPERVISOR"]
# Roles that can view admin dashboard (including MC_OFFICER with limited access)
ADMIN_VIEW_ROLES = ["ADMIN", "SUPERVISOR", "MC_OFFICER"]

# ============== PROPERTY ROUTES ==============

@api_router.get("/admin/properties")
async def list_properties(
    batch_id: Optional[str] = None,
    ward: Optional[str] = None,
    status: Optional[str] = None,
    employee_id: Optional[str] = None,
    search: Optional[str] = None,
    page: int = 1,
    limit: int = 50,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in ADMIN_VIEW_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    query = {}
    if batch_id and batch_id.strip():
        query["batch_id"] = batch_id
    if ward and ward.strip():
        query["ward"] = ward
    if status and status.strip():
        query["status"] = status
    if employee_id and employee_id.strip():
        query["assigned_employee_id"] = employee_id
    if search:
        query["$or"] = [
            {"property_id": {"$regex": search, "$options": "i"}},
            {"owner_name": {"$regex": search, "$options": "i"}},
            {"mobile": {"$regex": search, "$options": "i"}}
        ]
    
    skip = (page - 1) * limit
    total = await db.properties.count_documents(query)
    properties = await db.properties.find(query, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    
    return {
        "properties": properties,
        "total": total,
        "page": page,
        "pages": (total + limit - 1) // limit
    }

@api_router.post("/admin/assign")
async def assign_properties(data: AssignmentRequest, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    employee = await db.users.find_one({"id": data.employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    result = await db.properties.update_many(
        {"id": {"$in": data.property_ids}},
        {"$set": {
            "assigned_employee_id": data.employee_id,
            "assigned_employee_name": employee["name"]
        }}
    )
    
    return {"message": f"Assigned {result.modified_count} properties to {employee['name']}"}

@api_router.post("/admin/assign-bulk")
async def bulk_assign_by_ward(data: BulkAssignmentRequest, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "ADMIN":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    employee = await db.users.find_one({"id": data.employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    result = await db.properties.update_many(
        {"ward": data.area, "assigned_employee_id": None},
        {"$set": {
            "assigned_employee_id": data.employee_id,
            "assigned_employee_name": employee["name"]
        }}
    )
    
    await db.users.update_one(
        {"id": data.employee_id},
        {"$set": {"assigned_area": data.area}}
    )
    
    return {"message": f"Assigned {result.modified_count} properties in ward {data.area} to {employee['name']}"}

class BulkDeleteRequest(BaseModel):
    property_ids: List[str]

@api_router.post("/admin/properties/bulk-delete")
async def bulk_delete_properties(data: BulkDeleteRequest, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    if not data.property_ids:
        raise HTTPException(status_code=400, detail="No properties selected for deletion")
    
    # Delete associated submissions first
    await db.submissions.delete_many({"property_record_id": {"$in": data.property_ids}})
    
    # Delete the properties
    result = await db.properties.delete_many({"id": {"$in": data.property_ids}})
    
    return {
        "message": f"Successfully deleted {result.deleted_count} properties",
        "deleted_count": result.deleted_count
    }

@api_router.post("/admin/properties/delete-all")
async def delete_all_properties(
    batch_id: Optional[str] = None,
    ward: Optional[str] = None,
    status: Optional[str] = None,
    employee_id: Optional[str] = None,
    search: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Delete all properties matching the given filters. If no filters, deletes ALL properties."""
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Build query based on filters
    query = {}
    if batch_id and batch_id.strip():
        query["batch_id"] = batch_id
    if ward and ward.strip():
        query["ward"] = ward
    if status and status.strip():
        query["status"] = status
    if employee_id and employee_id.strip():
        query["assigned_employee_id"] = employee_id
    if search and search.strip():
        query["$or"] = [
            {"property_id": {"$regex": search, "$options": "i"}},
            {"owner_name": {"$regex": search, "$options": "i"}},
            {"mobile": {"$regex": search, "$options": "i"}}
        ]
    
    # Get count first
    count = await db.properties.count_documents(query)
    
    if count == 0:
        return {"message": "No properties found to delete", "deleted_count": 0}
    
    # Get all property IDs to delete submissions
    properties = await db.properties.find(query, {"id": 1, "_id": 0}).to_list(None)
    property_ids = [p["id"] for p in properties]
    
    # Delete associated submissions first
    await db.submissions.delete_many({"property_record_id": {"$in": property_ids}})
    
    # Delete the properties
    result = await db.properties.delete_many(query)
    
    return {
        "message": f"Successfully deleted {result.deleted_count} properties",
        "deleted_count": result.deleted_count
    }

@api_router.post("/admin/properties/arrange-by-route")
async def arrange_properties_by_route(
    ward: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Arrange properties by GPS route using nearest neighbor algorithm"""
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    query = {"latitude": {"$ne": None}, "longitude": {"$ne": None}}
    if ward and ward.strip():
        query["ward"] = ward
    
    # Get all properties with GPS
    properties = await db.properties.find(query, {"_id": 0}).to_list(None)
    
    if not properties:
        raise HTTPException(status_code=404, detail="No properties with GPS found")
    
    # Sort by GPS route using nearest neighbor algorithm
    sorted_props = []
    remaining = list(properties)
    
    if remaining:
        # Start from first property
        sorted_props.append(remaining.pop(0))
        
        while remaining:
            last = sorted_props[-1]
            last_lat, last_lon = last['latitude'], last['longitude']
            
            # Find nearest neighbor
            nearest_idx = 0
            nearest_dist = float('inf')
            
            for i, prop in enumerate(remaining):
                dist = haversine_distance(last_lat, last_lon, prop['latitude'], prop['longitude'])
                if dist < nearest_dist:
                    nearest_dist = dist
                    nearest_idx = i
            
            sorted_props.append(remaining.pop(nearest_idx))
    
    # Update serial numbers based on route order
    for i, prop in enumerate(sorted_props):
        await db.properties.update_one(
            {"id": prop["id"]},
            {"$set": {"serial_number": i + 1, "route_ordered": True}}
        )
    
    return {
        "message": f"Arranged {len(sorted_props)} properties by GPS route",
        "total_arranged": len(sorted_props)
    }

@api_router.post("/admin/properties/save-arranged")
async def save_arranged_data(
    ward: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Save the current arrangement as the permanent serial numbers"""
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    query = {}
    if ward and ward.strip():
        query["ward"] = ward
    
    # Get properties sorted by current serial_number
    properties = await db.properties.find(query, {"_id": 0}).sort("serial_number", 1).to_list(None)
    
    if not properties:
        raise HTTPException(status_code=404, detail="No properties found")
    
    # Re-assign serial numbers to ensure they're consecutive
    for i, prop in enumerate(properties):
        await db.properties.update_one(
            {"id": prop["id"]},
            {"$set": {"serial_number": i + 1, "arrangement_saved": True, "saved_at": datetime.now(timezone.utc).isoformat()}}
        )
    
    return {
        "message": f"Saved arrangement for {len(properties)} properties",
        "total_saved": len(properties)
    }

@api_router.post("/admin/properties/download-pdf")
async def download_properties_pdf(
    ward: Optional[str] = None,
    sn_position: str = "top-right",
    sn_font_size: int = 48,
    sn_color: str = "red",
    current_user: dict = Depends(get_current_user)
):
    """Generate PDF with property list arranged by serial number"""
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    query = {}
    if ward and ward.strip():
        query["ward"] = ward
    
    # Get properties sorted by serial_number
    properties = await db.properties.find(query, {"_id": 0}).sort("serial_number", 1).to_list(None)
    
    if not properties:
        raise HTTPException(status_code=404, detail="No properties found")
    
    # Generate PDF
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    colony_name = ward.replace(" ", "_") if ward else "all"
    pdf_filename = f"properties_{colony_name}_{timestamp}.pdf"
    pdf_path = UPLOAD_DIR / pdf_filename
    
    # Create PDF document
    doc = SimpleDocTemplate(str(pdf_path), pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        alignment=TA_CENTER,
        spaceAfter=20
    )
    elements.append(Paragraph(f"Property List - {ward or 'All Colonies'}", title_style))
    elements.append(Paragraph(f"Total: {len(properties)} properties | Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Table data
    table_data = [['SN', 'Property ID', 'Owner Name', 'Mobile', 'Category', 'Area', 'Amount']]
    
    for prop in properties:
        table_data.append([
            str(prop.get('serial_number', '-')),
            prop.get('property_id', '-'),
            prop.get('owner_name', '-')[:20] if prop.get('owner_name') else '-',
            prop.get('mobile', '-'),
            prop.get('category', '-')[:10] if prop.get('category') else '-',
            prop.get('total_area', '-'),
            f"₹{prop.get('amount', '0')}"
        ])
    
    # Create table
    table = Table(table_data, colWidths=[30, 70, 100, 80, 60, 50, 60])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    return {
        "message": f"Generated PDF with {len(properties)} properties",
        "filename": pdf_filename,
        "download_url": f"/api/uploads/{pdf_filename}"
    }

@api_router.get("/admin/wards")
async def list_wards(current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "ADMIN":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    wards = await db.properties.distinct("ward")
    wards = [w for w in wards if w]
    return {"wards": wards}

# ============== DASHBOARD ROUTES ==============

@api_router.get("/admin/dashboard", response_model=DashboardStats)
async def admin_dashboard(current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ADMIN_VIEW_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    today_start = get_today_start().isoformat()
    
    total = await db.properties.count_documents({})
    completed = await db.properties.count_documents({"status": "Completed"})
    pending = await db.properties.count_documents({"status": "Pending"})
    in_progress = await db.properties.count_documents({"status": "In Progress"})
    rejected = await db.properties.count_documents({"status": "Rejected"})
    employees = await db.users.count_documents({"role": {"$ne": "ADMIN"}})
    batches = await db.batches.count_documents({"status": "ACTIVE"})
    
    # Today's completed
    today_completed = await db.submissions.count_documents({
        "submitted_at": {"$gte": today_start},
        "status": {"$ne": "Rejected"}
    })
    
    # Today's unique wards (now called "colonies")
    today_submissions = await db.submissions.find(
        {"submitted_at": {"$gte": today_start}},
        {"property_record_id": 1, "_id": 0}
    ).to_list(10000)
    
    today_prop_ids = [s["property_record_id"] for s in today_submissions]
    if today_prop_ids:
        today_wards = await db.properties.distinct("ward", {"id": {"$in": today_prop_ids}})
        today_wards_count = len([w for w in today_wards if w])
    else:
        today_wards_count = 0
    
    return {
        "total_properties": total,
        "completed": completed,
        "pending": pending,
        "in_progress": in_progress,
        "rejected": rejected,
        "employees": employees,
        "batches": batches,
        "today_completed": today_completed,
        "today_wards": today_wards_count
    }

@api_router.get("/admin/employee-progress")
async def get_employee_progress(current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ADMIN_VIEW_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    today_start = get_today_start().isoformat()
    
    employees = await db.users.find({"role": {"$ne": "ADMIN"}}, {"_id": 0}).to_list(100)
    progress = []
    
    for emp in employees:
        total = await db.properties.count_documents({"assigned_employee_id": emp["id"]})
        completed = await db.properties.count_documents({
            "assigned_employee_id": emp["id"],
            "status": "Completed"
        })
        
        # Today's completed for this employee
        today_completed = await db.submissions.count_documents({
            "employee_id": emp["id"],
            "submitted_at": {"$gte": today_start},
            "status": {"$ne": "Rejected"}
        })
        
        # Overall completed (all time)
        overall_completed = await db.submissions.count_documents({
            "employee_id": emp["id"],
            "status": {"$ne": "Rejected"}
        })
        
        progress.append({
            "employee_id": emp["id"],
            "employee_name": emp["name"],
            "role": emp["role"],
            "total_assigned": total,
            "completed": completed,
            "pending": total - completed,
            "today_completed": today_completed,
            "overall_completed": overall_completed
        })
    
    return progress

# ============== SUBMISSIONS ROUTES ==============

@api_router.get("/admin/areas")
async def list_areas(current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Get unique areas/wards from properties
    areas = await db.properties.distinct("ward")
    # Filter out None/empty values and sort
    areas = sorted([a for a in areas if a])
    
    return {"areas": areas}

@api_router.get("/admin/submissions")
async def list_submissions(
    batch_id: Optional[str] = None,
    employee_id: Optional[str] = None,
    status: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    page: int = 1,
    limit: int = 50,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in ADMIN_VIEW_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    query = {}
    if batch_id and batch_id.strip():
        query["batch_id"] = batch_id
    if employee_id and employee_id.strip():
        query["employee_id"] = employee_id
    if status and status.strip():
        query["status"] = status
    if date_from:
        query["submitted_at"] = {"$gte": date_from}
    if date_to:
        if "submitted_at" in query:
            query["submitted_at"]["$lte"] = date_to
        else:
            query["submitted_at"] = {"$lte": date_to}
    
    skip = (page - 1) * limit
    total = await db.submissions.count_documents(query)
    submissions = await db.submissions.find(query, {"_id": 0}).sort("submitted_at", -1).skip(skip).limit(limit).to_list(limit)
    
    # Enrich with property details
    for sub in submissions:
        if sub.get("property_record_id"):
            prop = await db.properties.find_one({"id": sub["property_record_id"]}, {"_id": 0})
            if prop:
                sub["property_owner_name"] = prop.get("owner_name", "")
                sub["property_mobile"] = prop.get("mobile", "")
                sub["property_address"] = prop.get("address", "")
                sub["property_amount"] = prop.get("amount", "")
                sub["property_ward"] = prop.get("ward", "")
    
    return {
        "submissions": submissions,
        "total": total,
        "page": page,
        "pages": (total + limit - 1) // limit
    }

@api_router.post("/admin/submissions/approve")
async def approve_reject_submission(data: SubmissionApproval, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "ADMIN":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    submission = await db.submissions.find_one({"id": data.submission_id}, {"_id": 0})
    if not submission:
        raise HTTPException(status_code=404, detail="Submission not found")
    
    if data.action == "REJECT" and not data.remarks:
        raise HTTPException(status_code=400, detail="Remarks are required for rejection")
    
    new_status = "Approved" if data.action == "APPROVE" else "Rejected"
    
    update_data = {
        "status": new_status,
        "reviewed_by": current_user["id"],
        "reviewed_at": datetime.now(timezone.utc).isoformat()
    }
    
    if data.remarks:
        update_data["review_remarks"] = data.remarks
    
    await db.submissions.update_one(
        {"id": data.submission_id},
        {"$set": update_data}
    )
    
    # Update property status
    prop_status = "Completed" if data.action == "APPROVE" else "Rejected"
    await db.properties.update_one(
        {"id": submission["property_record_id"]},
        {"$set": {"status": prop_status}}
    )
    
    return {"message": f"Submission {new_status.lower()}"}

@api_router.put("/admin/submissions/{submission_id}")
async def edit_submission(
    submission_id: str,
    request: Request,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "ADMIN":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    data = await request.json()
    
    update_data = {}
    allowed_fields = [
        "new_owner_name", "new_mobile", "receiver_name", "relation",
        "old_property_id", "family_id", "aadhar_number", "ward_number",
        "remarks", "latitude", "longitude"
    ]
    
    for field in allowed_fields:
        if field in data:
            value = data[field]
            # Convert latitude/longitude to float if provided
            if field in ["latitude", "longitude"] and value:
                try:
                    update_data[field] = float(value)
                except ValueError:
                    pass
            else:
                update_data[field] = value
    
    update_data["edited_by"] = current_user["id"]
    update_data["edited_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.submissions.update_one(
        {"id": submission_id},
        {"$set": update_data}
    )
    
    return {"message": "Submission updated"}

@api_router.put("/admin/properties/{property_id}")
async def edit_property(
    property_id: str,
    request: Request,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "ADMIN":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    data = await request.json()
    
    update_data = {}
    allowed_fields = [
        "property_id", "owner_name", "mobile", "address", "amount", "ward"
    ]
    
    for field in allowed_fields:
        if field in data and data[field]:
            update_data[field] = data[field]
    
    update_data["edited_by"] = current_user["id"]
    update_data["edited_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.properties.update_one(
        {"id": property_id},
        {"$set": update_data}
    )
    
    return {"message": "Property updated"}

# ============== EXPORT ROUTES ==============

@api_router.get("/admin/export")
async def export_data(
    batch_id: Optional[str] = None,
    employee_id: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "ADMIN":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    query = {}
    if batch_id and batch_id.strip():
        query["batch_id"] = batch_id
    if employee_id and employee_id.strip():
        query["assigned_employee_id"] = employee_id
    if status and status.strip():
        query["status"] = status
    
    properties = await db.properties.find(query, {"_id": 0}).to_list(100000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Property Survey Data"
    
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    headers = [
        "Property ID", "Owner Name", "Mobile", "Address", "Total Area", "Amount", "Ward",
        "Assigned Employee", "Status", "New Owner Name", "New Mobile", "Receiver Name",
        "Relation", "Old Property ID", "Family ID", "Aadhar Number", "Ward Number",
        "GPS Latitude", "GPS Longitude", "Submission Date", "Signature URL", "Photo URLs",
        "Approval Status", "Review Remarks"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for row_idx, prop in enumerate(properties, 2):
        submission = await db.submissions.find_one(
            {"property_record_id": prop["id"]}, 
            {"_id": 0}
        )
        
        ws.cell(row=row_idx, column=1, value=prop.get("property_id", ""))
        ws.cell(row=row_idx, column=2, value=prop.get("owner_name", ""))
        ws.cell(row=row_idx, column=3, value=prop.get("mobile", ""))
        ws.cell(row=row_idx, column=4, value=prop.get("address", ""))
        ws.cell(row=row_idx, column=5, value=prop.get("total_area", ""))
        ws.cell(row=row_idx, column=6, value=prop.get("amount", ""))
        ws.cell(row=row_idx, column=7, value=prop.get("ward", ""))
        ws.cell(row=row_idx, column=8, value=prop.get("assigned_employee_name", ""))
        ws.cell(row=row_idx, column=9, value=prop.get("status", ""))
        
        if submission:
            ws.cell(row=row_idx, column=10, value=submission.get("new_owner_name", ""))
            ws.cell(row=row_idx, column=11, value=submission.get("new_mobile", ""))
            ws.cell(row=row_idx, column=12, value=submission.get("receiver_name", ""))
            ws.cell(row=row_idx, column=13, value=submission.get("relation", ""))
            ws.cell(row=row_idx, column=14, value=submission.get("old_property_id", ""))
            ws.cell(row=row_idx, column=15, value=submission.get("family_id", ""))
            ws.cell(row=row_idx, column=16, value=submission.get("aadhar_number", ""))
            ws.cell(row=row_idx, column=17, value=submission.get("ward_number", ""))
            ws.cell(row=row_idx, column=18, value=submission.get("latitude", ""))
            ws.cell(row=row_idx, column=19, value=submission.get("longitude", ""))
            ws.cell(row=row_idx, column=20, value=submission.get("submitted_at", ""))
            ws.cell(row=row_idx, column=21, value=submission.get("signature_url", ""))
            photos = submission.get("photos", [])
            photo_urls = ", ".join([p.get("file_url", "") for p in photos])
            ws.cell(row=row_idx, column=22, value=photo_urls)
            ws.cell(row=row_idx, column=23, value=submission.get("status", "Pending"))
            ws.cell(row=row_idx, column=24, value=submission.get("review_remarks", ""))
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    export_path = UPLOAD_DIR / f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(export_path)
    
    return FileResponse(
        path=str(export_path),
        filename=f"property_survey_export_{datetime.now().strftime('%Y%m%d')}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# Helper function to add watermark to photo
def add_watermark_to_photo(photo_path, latitude, longitude, submitted_at):
    try:
        img = PILImage.open(photo_path)
        draw = ImageDraw.Draw(img)
        
        if isinstance(submitted_at, str):
            try:
                dt = datetime.fromisoformat(submitted_at.replace('Z', '+00:00'))
            except:
                dt = datetime.now()
        else:
            dt = submitted_at or datetime.now()
        
        date_str = dt.strftime("%d/%m/%Y")
        time_str = dt.strftime("%I:%M:%S %p")
        
        watermark_lines = [
            f"Date: {date_str}",
            f"Time: {time_str}",
            f"Lat: {latitude:.6f}" if latitude else "Lat: N/A",
            f"Long: {longitude:.6f}" if longitude else "Long: N/A"
        ]
        
        font_size = max(16, min(img.width, img.height) // 25)
        try:
            font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", font_size)
        except:
            font = ImageFont.load_default()
        
        padding = font_size // 2
        line_height = font_size + 5
        
        max_text_width = max([draw.textlength(line, font=font) for line in watermark_lines])
        box_width = int(max_text_width + padding * 2)
        box_height = line_height * len(watermark_lines) + padding * 2
        
        box_x = padding
        box_y = img.height - box_height - padding
        
        overlay = PILImage.new('RGBA', img.size, (0, 0, 0, 0))
        overlay_draw = ImageDraw.Draw(overlay)
        overlay_draw.rectangle(
            [box_x, box_y, box_x + box_width, box_y + box_height],
            fill=(0, 0, 0, 180)
        )
        
        if img.mode != 'RGBA':
            img = img.convert('RGBA')
        
        img = PILImage.alpha_composite(img, overlay)
        draw = ImageDraw.Draw(img)
        
        for i, line in enumerate(watermark_lines):
            draw.text(
                (box_x + padding, box_y + padding + i * line_height),
                line,
                font=font,
                fill=(255, 255, 255, 255)
            )
        
        img = img.convert('RGB')
        
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.jpg')
        img.save(temp_file.name, 'JPEG', quality=90)
        return temp_file.name
    except Exception as e:
        logger.error(f"Error adding watermark: {e}")
        return photo_path

@api_router.get("/admin/export-pdf")
async def export_pdf(
    batch_id: Optional[str] = None,
    employee_id: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "ADMIN":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    query = {"status": "Completed"}
    if batch_id and batch_id.strip():
        query["batch_id"] = batch_id
    if employee_id and employee_id.strip():
        query["assigned_employee_id"] = employee_id
    if status and status.strip():
        query["status"] = status
    
    properties = await db.properties.find(query, {"_id": 0}).to_list(10000)
    
    pdf_path = UPLOAD_DIR / f"survey_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    doc = SimpleDocTemplate(
        str(pdf_path),
        pagesize=A4,
        rightMargin=20*mm,
        leftMargin=20*mm,
        topMargin=20*mm,
        bottomMargin=20*mm
    )
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=18, spaceAfter=20, alignment=TA_CENTER, textColor=colors.HexColor('#0f172a'))
    heading_style = ParagraphStyle('CustomHeading', parent=styles['Heading2'], fontSize=14, spaceAfter=10, textColor=colors.HexColor('#1e40af'))
    normal_style = ParagraphStyle('CustomNormal', parent=styles['Normal'], fontSize=10, spaceAfter=5)
    
    story = []
    
    story.append(Paragraph("NSTU Property Tax Survey Report", title_style))
    story.append(Paragraph(f"Generated on: {datetime.now().strftime('%d/%m/%Y %I:%M %p')}", normal_style))
    story.append(Spacer(1, 20))
    
    for prop in properties:
        submission = await db.submissions.find_one({"property_record_id": prop["id"]}, {"_id": 0})
        
        if not submission:
            continue
        
        story.append(Paragraph(f"Property ID: {prop.get('property_id', 'N/A')}", heading_style))
        
        prop_data = [
            ["Owner Name", prop.get("owner_name", "N/A")],
            ["Mobile", prop.get("mobile", "N/A")],
            ["Address", prop.get("address", "N/A")],
            ["Ward", prop.get("ward", "N/A")],
            ["Amount", prop.get("amount", "N/A")],
        ]
        
        prop_table = Table(prop_data, colWidths=[80*mm, 90*mm])
        prop_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f1f5f9')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#0f172a')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('PADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(prop_table)
        story.append(Spacer(1, 10))
        
        story.append(Paragraph("Survey Information", heading_style))
        survey_data = [
            ["New Owner Name", submission.get("new_owner_name", "N/A")],
            ["New Mobile", submission.get("new_mobile", "N/A")],
            ["Receiver Name", submission.get("receiver_name", "N/A")],
            ["Relation", submission.get("relation", "N/A")],
            ["Old Property ID", submission.get("old_property_id", "N/A")],
            ["Family ID", submission.get("family_id", "N/A")],
            ["Aadhar Number", submission.get("aadhar_number", "N/A")],
            ["Ward Number", submission.get("ward_number", "N/A")],
            ["Submitted By", submission.get("employee_name", "N/A")],
            ["Submitted At", submission.get("submitted_at", "N/A")],
            ["GPS Latitude", str(submission.get("latitude", "N/A"))],
            ["GPS Longitude", str(submission.get("longitude", "N/A"))],
            ["Status", submission.get("status", "Pending")],
        ]
        
        if submission.get("remarks"):
            survey_data.append(["Remarks", submission.get("remarks")])
        
        survey_table = Table(survey_data, colWidths=[80*mm, 90*mm])
        survey_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f1f5f9')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#0f172a')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('PADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(survey_table)
        story.append(Spacer(1, 15))
        
        photos = submission.get("photos", [])
        if photos:
            story.append(Paragraph("Photo Evidence (with GPS & Timestamp)", heading_style))
            
            for photo in photos:
                photo_url = photo.get("file_url", "")
                photo_type = photo.get("photo_type", "PHOTO")
                
                if photo_url.startswith("/api/uploads/"):
                    filename = photo_url.replace("/api/uploads/", "")
                    photo_path = UPLOAD_DIR / filename
                    
                    if photo_path.exists():
                        watermarked_path = add_watermark_to_photo(
                            str(photo_path),
                            submission.get("latitude"),
                            submission.get("longitude"),
                            submission.get("submitted_at")
                        )
                        
                        try:
                            img = RLImage(watermarked_path, width=80*mm, height=60*mm)
                            story.append(Paragraph(f"<b>{photo_type}</b>", normal_style))
                            story.append(img)
                            story.append(Spacer(1, 10))
                        except Exception as e:
                            logger.error(f"Error adding photo to PDF: {e}")
        
        signature_url = submission.get("signature_url")
        if signature_url:
            story.append(Paragraph("Property Holder Signature", heading_style))
            
            if signature_url.startswith("/api/uploads/"):
                sig_filename = signature_url.replace("/api/uploads/", "")
                sig_path = UPLOAD_DIR / sig_filename
                
                if sig_path.exists():
                    try:
                        sig_img = RLImage(str(sig_path), width=60*mm, height=30*mm)
                        sig_table = Table([[sig_img]], colWidths=[170*mm])
                        sig_table.setStyle(TableStyle([
                            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
                            ('BACKGROUND', (0, 0), (-1, -1), colors.white),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('PADDING', (0, 0), (-1, -1), 10),
                        ]))
                        story.append(sig_table)
                    except Exception as e:
                        logger.error(f"Error adding signature to PDF: {e}")
        
        story.append(PageBreak())
    
    doc.build(story)
    
    return FileResponse(
        path=str(pdf_path),
        filename=f"property_survey_report_{datetime.now().strftime('%Y%m%d')}.pdf",
        media_type="application/pdf"
    )

# ============== EMPLOYEE ROUTES ==============

@api_router.get("/employee/properties")
async def get_employee_properties(
    search: Optional[str] = None,
    status: Optional[str] = None,
    page: int = 1,
    limit: int = 20,
    current_user: dict = Depends(get_current_user)
):
    query = {"assigned_employee_id": current_user["id"]}
    if status and status.strip():
        query["status"] = status
    if search:
        query["$or"] = [
            {"property_id": {"$regex": search, "$options": "i"}},
            {"owner_name": {"$regex": search, "$options": "i"}},
            {"mobile": {"$regex": search, "$options": "i"}}
        ]
    
    skip = (page - 1) * limit
    total = await db.properties.count_documents(query)
    properties = await db.properties.find(query, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    
    return {
        "properties": properties,
        "total": total,
        "page": page,
        "pages": (total + limit - 1) // limit
    }

@api_router.get("/employee/property/{property_id}")
async def get_property_detail(property_id: str, current_user: dict = Depends(get_current_user)):
    prop = await db.properties.find_one({"id": property_id}, {"_id": 0})
    if not prop:
        raise HTTPException(status_code=404, detail="Property not found")
    
    if current_user["role"] != "ADMIN" and prop.get("assigned_employee_id") != current_user["id"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    submission = await db.submissions.find_one({"property_record_id": property_id}, {"_id": 0})
    
    return {
        "property": prop,
        "submission": submission
    }

@api_router.post("/employee/submit/{property_id}")
async def submit_survey(
    property_id: str,
    # Survey fields
    new_owner_name: str = Form(...),
    new_mobile: str = Form(...),
    receiver_name: str = Form(...),
    relation: str = Form(...),
    old_property_id: str = Form(None),
    family_id: str = Form(None),
    aadhar_number: str = Form(None),
    ward_number: str = Form(None),
    remarks: str = Form(None),
    self_satisfied: str = Form(None),  # New field: 'yes' or 'no'
    latitude: float = Form(...),
    longitude: float = Form(...),
    house_photo: UploadFile = File(...),
    gate_photo: UploadFile = File(...),
    signature: UploadFile = File(...),
    extra_photos: List[UploadFile] = File(default=[]),
    authorization: str = Form(...)
):
    current_user = await get_current_user(authorization)
    
    prop = await db.properties.find_one({"id": property_id}, {"_id": 0})
    if not prop:
        raise HTTPException(status_code=404, detail="Property not found")
    
    if current_user["role"] != "ADMIN" and prop.get("assigned_employee_id") != current_user["id"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    photos = []
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # House photo
    house_filename = f"{property_id}_house_{timestamp}{Path(house_photo.filename).suffix}"
    house_path = UPLOAD_DIR / house_filename
    async with aiofiles.open(house_path, 'wb') as f:
        content = await house_photo.read()
        await f.write(content)
    photos.append({"photo_type": "HOUSE", "file_url": f"/api/uploads/{house_filename}"})
    
    # Gate photo
    gate_filename = f"{property_id}_gate_{timestamp}{Path(gate_photo.filename).suffix}"
    gate_path = UPLOAD_DIR / gate_filename
    async with aiofiles.open(gate_path, 'wb') as f:
        content = await gate_photo.read()
        await f.write(content)
    photos.append({"photo_type": "GATE", "file_url": f"/api/uploads/{gate_filename}"})
    
    # Signature
    signature_filename = f"{property_id}_signature_{timestamp}.png"
    signature_path = UPLOAD_DIR / signature_filename
    async with aiofiles.open(signature_path, 'wb') as f:
        content = await signature.read()
        await f.write(content)
    signature_url = f"/api/uploads/{signature_filename}"
    
    # Extra photos
    for idx, photo in enumerate(extra_photos):
        if photo.filename:
            extra_filename = f"{property_id}_extra{idx}_{timestamp}{Path(photo.filename).suffix}"
            extra_path = UPLOAD_DIR / extra_filename
            async with aiofiles.open(extra_path, 'wb') as f:
                content = await photo.read()
                await f.write(content)
            photos.append({"photo_type": "EXTRA", "file_url": f"/api/uploads/{extra_filename}"})
    
    # Create submission with new fields
    submission_doc = {
        "id": str(uuid.uuid4()),
        "property_record_id": property_id,
        "property_id": prop["property_id"],
        "batch_id": prop["batch_id"],
        "employee_id": current_user["id"],
        "employee_name": current_user["name"],
        # Survey fields
        "new_owner_name": new_owner_name,
        "new_mobile": new_mobile,
        "receiver_name": receiver_name,
        "relation": relation,
        "old_property_id": old_property_id,
        "family_id": family_id,
        "aadhar_number": aadhar_number,
        "ward_number": ward_number,
        "remarks": remarks,
        "self_satisfied": self_satisfied,  # New field
        "latitude": latitude,
        "longitude": longitude,
        "submitted_at": datetime.now(timezone.utc).isoformat(),
        "photos": photos,
        "signature_url": signature_url,
        "status": "Pending"  # Pending approval
    }
    
    # Check if submission already exists
    existing = await db.submissions.find_one({"property_record_id": property_id})
    if existing:
        await db.submissions.update_one(
            {"property_record_id": property_id},
            {"$set": submission_doc}
        )
    else:
        await db.submissions.insert_one(submission_doc)
    
    # Update property status to In Progress (until approved)
    await db.properties.update_one(
        {"id": property_id},
        {"$set": {"status": "In Progress"}}
    )
    
    return {"message": "Survey submitted successfully", "submission_id": submission_doc["id"]}

@api_router.post("/employee/reject/{property_id}")
async def reject_property(property_id: str, remarks: str = Form(...), authorization: str = Form(...)):
    current_user = await get_current_user(authorization)
    
    prop = await db.properties.find_one({"id": property_id}, {"_id": 0})
    if not prop:
        raise HTTPException(status_code=404, detail="Property not found")
    
    if current_user["role"] != "ADMIN" and prop.get("assigned_employee_id") != current_user["id"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    await db.properties.update_one(
        {"id": property_id},
        {"$set": {"status": "Rejected", "reject_remarks": remarks}}
    )
    
    return {"message": "Property rejected"}

@api_router.get("/employee/progress")
async def get_employee_own_progress(current_user: dict = Depends(get_current_user)):
    today_start = get_today_start().isoformat()
    
    total = await db.properties.count_documents({"assigned_employee_id": current_user["id"]})
    completed = await db.properties.count_documents({
        "assigned_employee_id": current_user["id"],
        "status": "Completed"
    })
    pending = await db.properties.count_documents({
        "assigned_employee_id": current_user["id"],
        "status": "Pending"
    })
    rejected = await db.properties.count_documents({
        "assigned_employee_id": current_user["id"],
        "status": "Rejected"
    })
    in_progress = await db.properties.count_documents({
        "assigned_employee_id": current_user["id"],
        "status": "In Progress"
    })
    
    # Today's completed
    today_completed = await db.submissions.count_documents({
        "employee_id": current_user["id"],
        "submitted_at": {"$gte": today_start}
    })
    
    # Total completed (all time)
    total_completed = await db.submissions.count_documents({
        "employee_id": current_user["id"]
    })
    
    return {
        "total_assigned": total,
        "completed": completed,
        "pending": pending,
        "rejected": rejected,
        "in_progress": in_progress,
        "today_completed": today_completed,
        "total_completed": total_completed
    }

# ============== ATTENDANCE ROUTES ==============

@api_router.get("/employee/attendance/today")
async def check_today_attendance(current_user: dict = Depends(get_current_user)):
    """Check if employee has marked attendance today"""
    today_date = get_today_start().strftime("%Y-%m-%d")
    
    attendance = await db.attendance.find_one({
        "employee_id": current_user["id"],
        "date": today_date
    }, {"_id": 0})
    
    return {
        "has_attendance": attendance is not None,
        "attendance": attendance
    }

@api_router.post("/employee/attendance")
async def mark_attendance(
    selfie: UploadFile = File(...),
    latitude: float = Form(...),
    longitude: float = Form(...),
    authorization: str = Form(...)
):
    """Mark one-time daily attendance with selfie"""
    current_user = await get_current_user(authorization)
    today_date = get_today_start().strftime("%Y-%m-%d")
    
    # Check if already marked
    existing = await db.attendance.find_one({
        "employee_id": current_user["id"],
        "date": today_date
    })
    
    if existing:
        raise HTTPException(status_code=400, detail="Attendance already marked for today")
    
    # Save selfie
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    selfie_filename = f"attendance_{current_user['id']}_{timestamp}{Path(selfie.filename).suffix}"
    selfie_path = UPLOAD_DIR / selfie_filename
    async with aiofiles.open(selfie_path, 'wb') as f:
        content = await selfie.read()
        await f.write(content)
    
    selfie_url = f"/api/uploads/{selfie_filename}"
    
    # Create attendance record
    attendance_doc = {
        "id": str(uuid.uuid4()),
        "employee_id": current_user["id"],
        "employee_name": current_user["name"],
        "date": today_date,
        "marked_at": datetime.now(timezone.utc).isoformat(),
        "selfie_url": selfie_url,
        "latitude": latitude,
        "longitude": longitude
    }
    
    await db.attendance.insert_one(attendance_doc)
    
    return {
        "message": "Attendance marked successfully",
        "attendance_id": attendance_doc["id"],
        "marked_at": attendance_doc["marked_at"]
    }

@api_router.get("/admin/attendance")
async def get_attendance_records(
    date: Optional[str] = None,
    employee_id: Optional[str] = None,
    page: int = 1,
    limit: int = 50,
    current_user: dict = Depends(get_current_user)
):
    """Get attendance records (admin/supervisor only)"""
    if current_user["role"] not in ADMIN_VIEW_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    query = {}
    if date:
        query["date"] = date
    if employee_id and employee_id.strip():
        query["employee_id"] = employee_id
    
    skip = (page - 1) * limit
    total = await db.attendance.count_documents(query)
    records = await db.attendance.find(query, {"_id": 0}).sort("marked_at", -1).skip(skip).limit(limit).to_list(limit)
    
    return {
        "attendance": records,
        "total": total,
        "page": page,
        "pages": (total + limit - 1) // limit
    }

# ============== PDF BILL PROCESSING ==============

# Helper function to extract bill data from PDF text
def extract_bill_data(text: str, page_num: int) -> dict:
    """Extract structured bill data from PDF page text"""
    
    def find_value(patterns, text, default=""):
        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
            if match:
                return match.group(1).strip()
        return default
    
    # Extract coordinates (latitude : longitude format)
    coords_match = re.search(r'(\d+\.\d+)\s*:\s*(\d+\.\d+)', text)
    latitude = float(coords_match.group(1)) if coords_match else None
    longitude = float(coords_match.group(2)) if coords_match else None
    
    bill_data = {
        "bill_sr_no": find_value([r'Bill\s*Sr\s*No[:\s]*(\d+)', r'BillSrNo[:\s]*(\d+)'], text),
        "property_id": find_value([r'Property\s*Id[:\s]*([A-Z0-9]+)', r'PropertyId[:\s]*([A-Z0-9]+)'], text),
        "old_property_id": find_value([r'Old\s*Property\s*Id[:\s]*([A-Z0-9/-]+)', r'OldPropertyId[:\s]*([A-Z0-9/-]+)'], text),
        "financial_year": find_value([r'Financial\s*Year[:\s]*(\d{4}-\d{2,4})', r'FY[:\s]*(\d{4}-\d{2,4})'], text, "2025-26"),
        "print_date": find_value([r'Print\s*Date[:\s]*([0-9/\-]+)', r'Date[:\s]*([0-9/\-]+)'], text),
        "latitude": latitude,
        "longitude": longitude,
        "mobile": find_value([r'Mobile\s*No[:\s]*(\d{10})', r'Mobile[:\s]*(\d{10})', r'Phone[:\s]*(\d{10})'], text),
        "colony": find_value([r'Colony\s*Name[:\s]*([^\n]+)', r'Colony[:\s]*([^\n]+)'], text),
        "owner_name": find_value([r'Owner\s*Name[:\s]*([^\n]+)', r'Owner[:\s]*([^\n]+)'], text),
        "plot_address": find_value([r'Plot\s*Address[:\s]*([^\n]+)', r'Address[:\s]*([^\n]+)'], text),
        "permanent_address": find_value([r'Permanent\s*Address[:\s]*([^\n]+)'], text),
        "total_area": find_value([r'Total\s*Area[:\s]*([0-9.]+\s*SqYard)', r'Area[:\s]*([0-9.]+)'], text),
        "category": find_value([r'Category[:\s]*([^\n,]+)', r'Type[:\s]*([^\n,]+)'], text),
        "authorized_status": find_value([r'Authorized\s*Status[:\s]*([^\n]+)'], text),
        "total_outstanding": find_value([r'Total\s*Outstanding[:\s]*Rs?\.?\s*([0-9,.-]+)', r'Outstanding[:\s]*Rs?\.?\s*([0-9,.-]+)'], text),
        "property_tax_outstanding": find_value([r'Property\s*&?\s*Fire\s*Tax\s*Outstanding[:\s]*Rs?\.?\s*([0-9,.-]+)'], text),
        "page_number": page_num
    }
    
    return bill_data

# Calculate distance between two GPS points (Haversine formula)
def haversine_distance(lat1, lon1, lat2, lon2):
    R = 6371000  # Earth radius in meters
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    delta_phi = math.radians(lat2 - lat1)
    delta_lambda = math.radians(lon2 - lon1)
    a = math.sin(delta_phi/2)**2 + math.cos(phi1) * math.cos(phi2) * math.sin(delta_lambda/2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
    return R * c

# Sort bills by GPS route (nearest neighbor algorithm)
def sort_by_gps_route(bills: list) -> list:
    if not bills:
        return bills
    
    # Filter bills with valid GPS
    valid_bills = [b for b in bills if b.get('latitude') and b.get('longitude')]
    no_gps_bills = [b for b in bills if not b.get('latitude') or not b.get('longitude')]
    
    if not valid_bills:
        return bills
    
    # Start from the first bill
    sorted_bills = [valid_bills[0]]
    remaining = valid_bills[1:]
    
    while remaining:
        last = sorted_bills[-1]
        last_lat, last_lon = last['latitude'], last['longitude']
        
        # Find nearest neighbor
        nearest_idx = 0
        nearest_dist = float('inf')
        
        for i, bill in enumerate(remaining):
            dist = haversine_distance(last_lat, last_lon, bill['latitude'], bill['longitude'])
            if dist < nearest_dist:
                nearest_dist = dist
                nearest_idx = i
        
        sorted_bills.append(remaining.pop(nearest_idx))
    
    # Add bills without GPS at the end
    sorted_bills.extend(no_gps_bills)
    
    return sorted_bills

@api_router.post("/admin/bills/upload-pdf")
async def upload_pdf_bills(
    file: UploadFile = File(...),
    batch_name: str = Form(...),
    authorization: str = Form(...)
):
    """Upload multi-page PDF and extract bill data from each page"""
    current_user = await get_current_user(authorization)
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Please upload a PDF file")
    
    # Save uploaded PDF
    content = await file.read()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    pdf_filename = f"bills_{timestamp}.pdf"
    pdf_path = UPLOAD_DIR / pdf_filename
    
    async with aiofiles.open(pdf_path, 'wb') as f:
        await f.write(content)
    
    # Create batch record
    batch_id = str(uuid.uuid4())
    batch_doc = {
        "id": batch_id,
        "name": batch_name,
        "type": "PDF_BILLS",
        "pdf_filename": pdf_filename,
        "pdf_url": f"/api/uploads/{pdf_filename}",
        "uploaded_by": current_user["id"],
        "uploaded_at": datetime.now(timezone.utc).isoformat(),
        "status": "ACTIVE",
        "total_records": 0
    }
    
    # Extract text from each page using PyMuPDF
    bills = []
    try:
        pdf_doc = fitz.open(str(pdf_path))
        
        for page_num in range(len(pdf_doc)):
            page = pdf_doc[page_num]
            text = page.get_text()
            
            # Extract bill data
            bill_data = extract_bill_data(text, page_num + 1)
            bill_data["id"] = str(uuid.uuid4())
            bill_data["batch_id"] = batch_id
            bill_data["serial_number"] = page_num + 1  # Initial serial number
            bill_data["created_at"] = datetime.now(timezone.utc).isoformat()
            bill_data["status"] = "Pending"
            
            bills.append(bill_data)
        
        pdf_doc.close()
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing PDF: {str(e)}")
    
    # Insert bills into database
    if bills:
        await db.bills.insert_many(bills)
        batch_doc["total_records"] = len(bills)
    
    await db.batches.insert_one(batch_doc)
    
    # Get unique colonies
    colonies = list(set([b.get("colony", "").strip() for b in bills if b.get("colony")]))
    
    return {
        "batch_id": batch_id,
        "name": batch_name,
        "total_bills": len(bills),
        "colonies": colonies,
        "message": f"Successfully extracted {len(bills)} bills from PDF"
    }

@api_router.get("/admin/bills")
async def list_bills(
    batch_id: Optional[str] = None,
    colony: Optional[str] = None,
    status: Optional[str] = None,
    sorted_by_route: bool = False,
    page: int = 1,
    limit: int = 50,
    current_user: dict = Depends(get_current_user)
):
    """Get bills with optional filtering"""
    if current_user["role"] not in ADMIN_VIEW_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    query = {}
    if batch_id and batch_id.strip():
        query["batch_id"] = batch_id
    if colony and colony.strip():
        query["colony"] = {"$regex": colony, "$options": "i"}
    if status and status.strip():
        query["status"] = status
    
    total = await db.bills.count_documents(query)
    
    if sorted_by_route:
        # Get all matching bills and sort by GPS route
        all_bills = await db.bills.find(query, {"_id": 0}).to_list(None)
        sorted_bills = sort_by_gps_route(all_bills)
        
        # Assign new serial numbers
        for i, bill in enumerate(sorted_bills):
            bill["route_serial"] = i + 1
        
        # Paginate
        start = (page - 1) * limit
        bills = sorted_bills[start:start + limit]
    else:
        bills = await db.bills.find(query, {"_id": 0}).sort("serial_number", 1).skip((page - 1) * limit).limit(limit).to_list(limit)
    
    return {
        "bills": bills,
        "total": total,
        "page": page,
        "pages": (total + limit - 1) // limit
    }

@api_router.get("/admin/bills/colonies")
async def get_bill_colonies(
    batch_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get unique colonies from bills"""
    if current_user["role"] not in ADMIN_VIEW_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    query = {}
    if batch_id and batch_id.strip():
        query["batch_id"] = batch_id
    
    colonies = await db.bills.distinct("colony", query)
    colonies = [c for c in colonies if c and c.strip()]
    
    return {"colonies": sorted(colonies)}

@api_router.put("/admin/bills/{bill_id}")
async def update_bill(
    bill_id: str,
    current_user: dict = Depends(get_current_user),
    owner_name: str = Form(None),
    mobile: str = Form(None),
    plot_address: str = Form(None),
    permanent_address: str = Form(None),
    category: str = Form(None),
    total_area: str = Form(None),
    total_outstanding: str = Form(None),
    colony: str = Form(None)
):
    """Edit bill data"""
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    bill = await db.bills.find_one({"id": bill_id})
    if not bill:
        raise HTTPException(status_code=404, detail="Bill not found")
    
    update_data = {"updated_at": datetime.now(timezone.utc).isoformat()}
    if owner_name is not None:
        update_data["owner_name"] = owner_name
    if mobile is not None:
        update_data["mobile"] = mobile
    if plot_address is not None:
        update_data["plot_address"] = plot_address
    if permanent_address is not None:
        update_data["permanent_address"] = permanent_address
    if category is not None:
        update_data["category"] = category
    if total_area is not None:
        update_data["total_area"] = total_area
    if total_outstanding is not None:
        update_data["total_outstanding"] = total_outstanding
    if colony is not None:
        update_data["colony"] = colony
    
    await db.bills.update_one({"id": bill_id}, {"$set": update_data})
    
    return {"message": "Bill updated successfully"}

@api_router.post("/admin/bills/arrange-by-route")
async def arrange_bills_by_route(
    batch_id: str = Form(None),
    colony: str = Form(None),
    current_user: dict = Depends(get_current_user)
):
    """Arrange bills by GPS route and assign new serial numbers"""
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    query = {}
    if batch_id and batch_id.strip():
        query["batch_id"] = batch_id
    if colony and colony.strip():
        query["colony"] = {"$regex": colony, "$options": "i"}
    
    # Get all matching bills
    bills = await db.bills.find(query, {"_id": 0}).to_list(None)
    
    if not bills:
        raise HTTPException(status_code=404, detail="No bills found")
    
    # Sort by GPS route
    sorted_bills = sort_by_gps_route(bills)
    
    # Update serial numbers in database
    for i, bill in enumerate(sorted_bills):
        await db.bills.update_one(
            {"id": bill["id"]},
            {"$set": {"serial_number": i + 1, "route_ordered": True}}
        )
    
    return {
        "message": f"Arranged {len(sorted_bills)} bills by GPS route",
        "total_arranged": len(sorted_bills)
    }

@api_router.post("/admin/bills/generate-pdf")
async def generate_arranged_pdf(
    batch_id: str = Form(None),
    colony: str = Form(None),
    sn_position: str = Form("top-right"),  # top-left, top-right, bottom-left, bottom-right
    sn_font_size: int = Form(48),
    sn_color: str = Form("red"),
    current_user: dict = Depends(get_current_user)
):
    """Generate arranged PDF with serial numbers"""
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    query = {}
    if batch_id and batch_id.strip():
        query["batch_id"] = batch_id
    if colony and colony.strip():
        query["colony"] = {"$regex": colony, "$options": "i"}
    
    # Get arranged bills
    bills = await db.bills.find(query, {"_id": 0}).sort("serial_number", 1).to_list(None)
    
    if not bills:
        raise HTTPException(status_code=404, detail="No bills found")
    
    # Get original PDF
    batch = await db.batches.find_one({"id": bills[0]["batch_id"]})
    if not batch or not batch.get("pdf_filename"):
        raise HTTPException(status_code=404, detail="Original PDF not found")
    
    original_pdf_path = UPLOAD_DIR / batch["pdf_filename"]
    if not original_pdf_path.exists():
        raise HTTPException(status_code=404, detail="Original PDF file not found")
    
    # Create new PDF with serial numbers
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f"arranged_{colony or 'all'}_{timestamp}.pdf"
    output_path = UPLOAD_DIR / output_filename
    
    # Color mapping
    color_map = {
        "red": (1, 0, 0),
        "blue": (0, 0, 1),
        "green": (0, 0.5, 0),
        "black": (0, 0, 0),
        "orange": (1, 0.5, 0)
    }
    sn_rgb = color_map.get(sn_color.lower(), (1, 0, 0))
    
    # Open original PDF and create new one with SN overlay
    src_pdf = fitz.open(str(original_pdf_path))
    output_pdf = fitz.open()
    
    for bill in bills:
        page_num = bill.get("page_number", 1) - 1
        if page_num < 0 or page_num >= len(src_pdf):
            continue
        
        # Copy page
        output_pdf.insert_pdf(src_pdf, from_page=page_num, to_page=page_num)
        new_page = output_pdf[-1]
        
        # Calculate SN position - adjusted for BillSrNo field location
        rect = new_page.rect
        
        # BillSrNo field is typically at top-right of the bill
        # Standard A4: 595 x 842 points
        if sn_position == "top-left":
            x, y = 50, 60
        elif sn_position == "top-right":
            # Position near the BillSrNo.: field area (top-right of bill)
            x, y = rect.width - 80, 45
        elif sn_position == "bottom-left":
            x, y = 50, rect.height - 50
        else:  # bottom-right
            x, y = rect.width - 80, rect.height - 50
        
        # Add serial number text (plain number: 1, 2, 3...)
        sn_text = f"{bill['serial_number']}"
        new_page.insert_text(
            (x, y),
            sn_text,
            fontsize=sn_font_size,
            color=sn_rgb,
            fontname="helv"
        )
    
    output_pdf.save(str(output_path))
    output_pdf.close()
    src_pdf.close()
    
    return {
        "message": f"Generated PDF with {len(bills)} bills",
        "filename": output_filename,
        "download_url": f"/api/uploads/{output_filename}"
    }

@api_router.post("/admin/bills/split-by-employee")
async def split_bills_by_employee(
    batch_id: str = Form(None),
    colony: str = Form(None),
    employee_count: int = Form(...),
    sn_font_size: int = Form(48),
    sn_color: str = Form("red"),
    current_user: dict = Depends(get_current_user)
):
    """Split bills into separate PDFs for each employee"""
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    if employee_count < 1 or employee_count > 100:
        raise HTTPException(status_code=400, detail="Employee count must be between 1 and 100")
    
    query = {}
    if batch_id and batch_id.strip():
        query["batch_id"] = batch_id
    if colony and colony.strip():
        query["colony"] = {"$regex": colony, "$options": "i"}
    
    # Get arranged bills
    bills = await db.bills.find(query, {"_id": 0}).sort("serial_number", 1).to_list(None)
    
    if not bills:
        raise HTTPException(status_code=404, detail="No bills found")
    
    # Get original PDF
    batch = await db.batches.find_one({"id": bills[0]["batch_id"]})
    if not batch or not batch.get("pdf_filename"):
        raise HTTPException(status_code=404, detail="Original PDF not found")
    
    original_pdf_path = UPLOAD_DIR / batch["pdf_filename"]
    if not original_pdf_path.exists():
        raise HTTPException(status_code=404, detail="Original PDF file not found")
    
    # Calculate bills per employee
    total_bills = len(bills)
    bills_per_employee = math.ceil(total_bills / employee_count)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    generated_files = []
    
    # Color mapping
    color_map = {
        "red": (1, 0, 0),
        "blue": (0, 0, 1),
        "green": (0, 0.5, 0),
        "black": (0, 0, 0),
        "orange": (1, 0.5, 0)
    }
    sn_rgb = color_map.get(sn_color.lower(), (1, 0, 0))
    
    src_pdf = fitz.open(str(original_pdf_path))
    
    for emp_idx in range(employee_count):
        start_idx = emp_idx * bills_per_employee
        end_idx = min(start_idx + bills_per_employee, total_bills)
        
        if start_idx >= total_bills:
            break
        
        employee_bills = bills[start_idx:end_idx]
        
        output_filename = f"employee_{emp_idx + 1}_{colony or 'all'}_{timestamp}.pdf"
        output_path = UPLOAD_DIR / output_filename
        
        output_pdf = fitz.open()
        
        for bill in employee_bills:
            page_num = bill.get("page_number", 1) - 1
            if page_num < 0 or page_num >= len(src_pdf):
                continue
            
            output_pdf.insert_pdf(src_pdf, from_page=page_num, to_page=page_num)
            new_page = output_pdf[-1]
            
            # Add serial number to top-right - near BillSrNo field
            rect = new_page.rect
            x, y = rect.width - 80, 45
            sn_text = f"{bill['serial_number']}"
            new_page.insert_text((x, y), sn_text, fontsize=sn_font_size, color=sn_rgb, fontname="helv")
        
        output_pdf.save(str(output_path))
        output_pdf.close()
        
        generated_files.append({
            "employee_number": emp_idx + 1,
            "filename": output_filename,
            "download_url": f"/api/uploads/{output_filename}",
            "bill_range": f"SN {employee_bills[0]['serial_number']} - {employee_bills[-1]['serial_number']}",
            "total_bills": len(employee_bills)
        })
    
    src_pdf.close()
    
    return {
        "message": f"Generated {len(generated_files)} employee PDFs",
        "total_bills": total_bills,
        "bills_per_employee": bills_per_employee,
        "files": generated_files
    }

@api_router.get("/admin/bills/map-data")
async def get_bills_map_data(
    batch_id: Optional[str] = None,
    colony: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get bill data for map display"""
    if current_user["role"] not in ADMIN_VIEW_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    query = {}
    if batch_id and batch_id.strip():
        query["batch_id"] = batch_id
    if colony and colony.strip():
        query["colony"] = {"$regex": colony, "$options": "i"}
    
    # Only get bills with GPS coordinates
    query["latitude"] = {"$ne": None}
    query["longitude"] = {"$ne": None}
    
    bills = await db.bills.find(query, {
        "_id": 0,
        "id": 1,
        "serial_number": 1,
        "property_id": 1,
        "owner_name": 1,
        "mobile": 1,
        "colony": 1,
        "latitude": 1,
        "longitude": 1,
        "total_outstanding": 1,
        "category": 1
    }).sort("serial_number", 1).to_list(None)
    
    return {
        "bills": bills,
        "total": len(bills)
    }

@api_router.delete("/admin/bills/batch/{batch_id}")
async def delete_bill_batch(batch_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a bill batch and all its bills"""
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Delete bills
    result = await db.bills.delete_many({"batch_id": batch_id})
    
    # Delete batch
    await db.batches.delete_one({"id": batch_id})
    
    return {"message": f"Deleted batch and {result.deleted_count} bills"}

@api_router.post("/admin/bills/delete-all")
async def delete_all_bills(
    batch_id: str = Form(None),
    colony: str = Form(None),
    current_user: dict = Depends(get_current_user)
):
    """Delete all bills matching the given filters. If no filters, deletes ALL bills."""
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    query = {}
    if batch_id and batch_id.strip():
        query["batch_id"] = batch_id
    if colony and colony.strip():
        query["colony"] = {"$regex": colony, "$options": "i"}
    
    # Get count first
    count = await db.bills.count_documents(query)
    
    if count == 0:
        return {"message": "No bills found to delete", "deleted_count": 0}
    
    # Delete the bills
    result = await db.bills.delete_many(query)
    
    # Update batch record counts if batch_id specified
    if batch_id and batch_id.strip():
        remaining = await db.bills.count_documents({"batch_id": batch_id})
        await db.batches.update_one(
            {"id": batch_id},
            {"$set": {"total_records": remaining}}
        )
        # If no bills left, delete the batch
        if remaining == 0:
            await db.batches.delete_one({"id": batch_id})
    
    return {
        "message": f"Successfully deleted {result.deleted_count} bills",
        "deleted_count": result.deleted_count
    }

@api_router.post("/admin/bills/copy-to-properties")
async def copy_bills_to_properties(
    batch_id: str = Form(None),
    colony: str = Form(None),
    current_user: dict = Depends(get_current_user)
):
    """Copy bill data to properties collection"""
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    query = {}
    if batch_id and batch_id.strip():
        query["batch_id"] = batch_id
    if colony and colony.strip():
        query["colony"] = {"$regex": colony, "$options": "i"}
    
    # Get bills to copy
    bills = await db.bills.find(query, {"_id": 0}).sort("serial_number", 1).to_list(None)
    
    if not bills:
        raise HTTPException(status_code=404, detail="No bills found to copy")
    
    # Create a new batch for properties
    prop_batch_id = str(uuid.uuid4())
    prop_batch_name = f"Bills Import {colony or 'All'} - {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    
    prop_batch_doc = {
        "id": prop_batch_id,
        "name": prop_batch_name,
        "uploaded_by": current_user["id"],
        "uploaded_at": datetime.now(timezone.utc).isoformat(),
        "status": "ACTIVE",
        "total_records": len(bills),
        "source": "PDF_BILLS"
    }
    await db.batches.insert_one(prop_batch_doc)
    
    # Convert bills to properties
    properties = []
    for i, bill in enumerate(bills):
        prop = {
            "id": str(uuid.uuid4()),
            "batch_id": prop_batch_id,
            "serial_number": i + 1,
            "property_id": bill.get("property_id", str(uuid.uuid4())[:8].upper()),
            "old_property_id": bill.get("old_property_id", ""),
            "owner_name": bill.get("owner_name", "Unknown"),
            "mobile": bill.get("mobile", ""),
            "address": bill.get("plot_address", ""),
            "colony": bill.get("colony", ""),
            "ward": bill.get("colony", ""),  # Use colony as ward
            "latitude": bill.get("latitude"),
            "longitude": bill.get("longitude"),
            "total_area": bill.get("total_area", ""),
            "category": bill.get("category", ""),
            "amount": bill.get("total_outstanding", "0"),
            "financial_year": bill.get("financial_year", "2025-2026"),
            "assigned_employee_id": None,
            "assigned_employee_name": None,
            "status": "Pending",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "source_bill_id": bill.get("id")  # Reference to original bill
        }
        properties.append(prop)
    
    # Insert properties
    if properties:
        await db.properties.insert_many(properties)
    
    return {
        "message": f"Successfully added {len(properties)} bills to properties",
        "batch_id": prop_batch_id,
        "batch_name": prop_batch_name,
        "total_added": len(properties)
    }

@api_router.post("/admin/bills/split-by-employees")
async def split_bills_by_specific_employees(
    batch_id: str = Form(None),
    colony: str = Form(None),
    employee_ids: str = Form(...),  # Comma-separated employee IDs
    sn_font_size: int = Form(48),
    sn_color: str = Form("red"),
    current_user: dict = Depends(get_current_user)
):
    """Split bills among specific employees and generate separate PDFs"""
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Parse employee IDs
    emp_ids = [e.strip() for e in employee_ids.split(",") if e.strip()]
    
    if not emp_ids:
        raise HTTPException(status_code=400, detail="At least one employee must be selected")
    
    # Verify employees exist
    employees = []
    for emp_id in emp_ids:
        emp = await db.users.find_one({"id": emp_id}, {"_id": 0, "id": 1, "name": 1, "username": 1})
        if emp:
            employees.append(emp)
    
    if not employees:
        raise HTTPException(status_code=404, detail="No valid employees found")
    
    query = {}
    if batch_id and batch_id.strip():
        query["batch_id"] = batch_id
    if colony and colony.strip():
        query["colony"] = {"$regex": colony, "$options": "i"}
    
    # Get arranged bills
    bills = await db.bills.find(query, {"_id": 0}).sort("serial_number", 1).to_list(None)
    
    if not bills:
        raise HTTPException(status_code=404, detail="No bills found")
    
    # Get original PDF
    batch = await db.batches.find_one({"id": bills[0]["batch_id"]})
    if not batch or not batch.get("pdf_filename"):
        raise HTTPException(status_code=404, detail="Original PDF not found")
    
    original_pdf_path = UPLOAD_DIR / batch["pdf_filename"]
    if not original_pdf_path.exists():
        raise HTTPException(status_code=404, detail="Original PDF file not found")
    
    # Calculate bills per employee
    total_bills = len(bills)
    employee_count = len(employees)
    bills_per_employee = math.ceil(total_bills / employee_count)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    generated_files = []
    
    # Color mapping
    color_map = {
        "red": (1, 0, 0),
        "blue": (0, 0, 1),
        "green": (0, 0.5, 0),
        "black": (0, 0, 0),
        "orange": (1, 0.5, 0)
    }
    sn_rgb = color_map.get(sn_color.lower(), (1, 0, 0))
    
    src_pdf = fitz.open(str(original_pdf_path))
    
    for emp_idx, emp in enumerate(employees):
        start_idx = emp_idx * bills_per_employee
        end_idx = min(start_idx + bills_per_employee, total_bills)
        
        if start_idx >= total_bills:
            break
        
        employee_bills = bills[start_idx:end_idx]
        
        # Use employee name in filename (sanitize for filename)
        emp_name_safe = re.sub(r'[^\w\-_]', '_', emp.get('name', f'emp_{emp_idx+1}'))
        output_filename = f"{emp_name_safe}_{colony or 'all'}_{timestamp}.pdf"
        output_path = UPLOAD_DIR / output_filename
        
        output_pdf = fitz.open()
        
        for bill in employee_bills:
            page_num = bill.get("page_number", 1) - 1
            if page_num < 0 or page_num >= len(src_pdf):
                continue
            
            output_pdf.insert_pdf(src_pdf, from_page=page_num, to_page=page_num)
            new_page = output_pdf[-1]
            
            # Add serial number to top-right - near BillSrNo field
            rect = new_page.rect
            x, y = rect.width - 80, 45
            sn_text = f"{bill['serial_number']}"
            new_page.insert_text((x, y), sn_text, fontsize=sn_font_size, color=sn_rgb, fontname="helv")
        
        output_pdf.save(str(output_path))
        output_pdf.close()
        
        generated_files.append({
            "employee_id": emp["id"],
            "employee_name": emp.get("name", emp.get("username", f"Employee {emp_idx+1}")),
            "filename": output_filename,
            "download_url": f"/api/uploads/{output_filename}",
            "bill_range": f"SR {employee_bills[0]['serial_number']} - {employee_bills[-1]['serial_number']}",
            "total_bills": len(employee_bills)
        })
    
    src_pdf.close()
    
    return {
        "message": f"Generated PDFs for {len(generated_files)} employees",
        "total_bills": total_bills,
        "bills_per_employee": bills_per_employee,
        "files": generated_files
    }

# ============== FILE SERVING ==============

@api_router.get("/uploads/{filename}")
async def serve_upload(filename: str):
    file_path = UPLOAD_DIR / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(str(file_path))

# ============== INITIALIZATION ==============

@api_router.get("/")
async def root():
    return {"message": "NSTU Property Tax Manager API"}

@api_router.post("/init-admin")
async def init_admin():
    """Initialize default admin user if not exists"""
    existing = await db.users.find_one({"username": "admin"})
    if existing:
        return {"message": "Admin already exists"}
    
    admin_doc = {
        "id": str(uuid.uuid4()),
        "username": "admin",
        "password_hash": hash_password("admin123"),
        "name": "Super Admin",
        "role": "ADMIN",
        "assigned_area": None,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(admin_doc)
    return {"message": "Admin user created", "username": "admin", "password": "admin123"}

# Include the router
app.include_router(api_router)

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
